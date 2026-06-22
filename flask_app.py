import os
import json
import logging
import traceback
from io import BytesIO
from datetime import date, timedelta, datetime
from flask import (Flask, render_template, request, session,
                   redirect, url_for, flash, send_file)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text

app = Flask(__name__)
app.secret_key = 'tabor-tajny-klic-2026-zmente-me'

basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = (
    'sqlite:///' + os.path.join(basedir, 'tabor.db'))
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    app.logger.error('500 chyba:\n' + traceback.format_exc())
    return render_template('500.html'), 500

# ── Paleta barev (Word-like) ─────────────────────────────────────────────────
PALETTE_COLORS = [
    '#FF5252','#D32F2F','#FF7043','#FF9800','#FFC107',
    '#FFEE58','#C6E03A','#4CAF50','#388E3C','#00897B',
    '#26C6DA','#039BE5','#1E88E5','#1565C0','#3949AB',
    '#7B1FA2','#E91E63','#AD1457','#795548','#546E7A',
]

MEAL_LABELS = {
    'snidane':    'Snídaně',
    'dop_svacina':'Dop. svačina',
    'obed':       'Oběd',
    'odp_svacina':'Odp. svačina',
    'vecere':     'Večeře',
}
MEAL_TYPES = list(MEAL_LABELS.keys())

INTENSITY_LABELS = {
    'nizka':  'Nízká',
    'stredni':'Střední',
    'vysoka': 'Vysoká',
}

INTENSITY_ICONS = {
    'nizka':   'bi-lungs-fill',
    'stredni': 'bi-lungs-fill',
    'vysoka':  'bi-lungs-fill',
}

GAME_TYPE_LABELS = {
    'hra':      'Táborová hra',
    'aktivita': 'Celotáborová aktivita',
}
GAME_TYPE_ICONS = {
    'hra':      'bi-dice-5-fill',
    'aktivita': 'bi-fire',
}
GAME_TYPE_COLORS = {
    'hra':      {'bg': '#e8f4fd', 'border': '#90caf9', 'text': 'text-primary'},
    'aktivita': {'bg': '#e8f5e9', 'border': '#a5d6a7', 'text': 'text-success'},
}

CHALLENGE_NAMES = {
    'orli_pera':   'Orlí pera',
    'kapky_rosy':  'Kapky rosy',
    'bile_tesaky': 'Bílé tesáky',
}
CHALLENGE_AREAS = {
    'sever': 'Severně od tábora',
    'zapad': 'Západně od tábora',
    'jih':   'Jižně od tábora',
}
CHALLENGE_COLORS = {
    'orli_pera':   '#D4845A',
    'kapky_rosy':  '#6A9FC0',
    'bile_tesaky': '#7A909A',
}
AREA_ICONS = {
    'sever': '↑ (severně od tábora)',
    'zapad': '← (západně od tábora)',
    'jih':   '↓ (jižně od tábora)',
}

# ── Modely ───────────────────────────────────────────────────────────────────

class Camp(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(200), nullable=False)
    date_from  = db.Column(db.Date, nullable=False)
    date_to    = db.Column(db.Date, nullable=False)
    troops         = db.relationship('Troop', backref='camp', lazy=True,
                                     cascade='all, delete-orphan')
    slot_templates = db.relationship('SlotTemplate', backref='camp', lazy=True,
                                     order_by='SlotTemplate.order',
                                     cascade='all, delete-orphan')
    service_slots  = db.relationship('ServiceSlotTemplate', backref='camp',
                                     lazy=True, cascade='all, delete-orphan')

    def get_days(self):
        days, d = [], self.date_from
        while d <= self.date_to:
            days.append(d)
            d += timedelta(days=1)
        return days


class Troop(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    name    = db.Column(db.String(100), nullable=False)
    color   = db.Column(db.String(7), nullable=False, default='#4CAF50')
    order   = db.Column(db.Integer, nullable=False, default=0)
    camp_id = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)


class SlotTemplate(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    name      = db.Column(db.String(100), nullable=False)
    time_from = db.Column(db.String(5), nullable=False)
    time_to   = db.Column(db.String(5), nullable=False)
    order     = db.Column(db.Integer, nullable=False, default=0)
    camp_id   = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)


class ServiceSlotTemplate(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    slot_number = db.Column(db.Integer, nullable=False)
    time_from   = db.Column(db.String(5), nullable=False, default='00:00')
    time_to     = db.Column(db.String(5), nullable=False, default='12:00')
    camp_id     = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)


class ProgramEntry(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    camp_id   = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)
    troop_id  = db.Column(db.Integer, db.ForeignKey('troop.id'), nullable=False)
    date      = db.Column(db.Date, nullable=False)
    slot_ids        = db.Column(db.String(200), nullable=False)
    time_from       = db.Column(db.String(5))
    time_to         = db.Column(db.String(5))
    title           = db.Column(db.String(200))
    shared_group_id = db.Column(db.Integer, nullable=True)
    teepee          = db.Column(db.Boolean, default=False, nullable=True)
    troop           = db.relationship('Troop', backref='program_entries')

    def get_slot_ids(self):
        return json.loads(self.slot_ids) if self.slot_ids else []


class ProgramItem(db.Model):
    """Pomůcka / materiál pro programový blok oddílu."""
    id       = db.Column(db.Integer, primary_key=True)
    camp_id  = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)
    entry_id = db.Column(db.Integer, db.ForeignKey('program_entry.id'), nullable=False)
    name     = db.Column(db.String(300), nullable=False)
    checked  = db.Column(db.Boolean, default=False, nullable=False)
    entry    = db.relationship('ProgramEntry', backref='items')


class ActivityLog(db.Model):
    """Log posledních úprav v programu tábora."""
    id          = db.Column(db.Integer, primary_key=True)
    camp_id     = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)
    timestamp   = db.Column(db.DateTime, nullable=False)
    category    = db.Column(db.String(30), nullable=False)   # program/hra/aktivita/meals/service/challenge/item
    description = db.Column(db.String(400), nullable=False)
    link_url    = db.Column(db.String(300), nullable=True)

    @property
    def timestamp_display(self):
        now = datetime.now()
        delta = (now.date() - self.timestamp.date()).days
        t = self.timestamp.strftime('%H:%M')
        if delta == 0:   return f'dnes {t}'
        if delta == 1:   return f'včera {t}'
        return self.timestamp.strftime('%-d. %-m.') + f' {t}'


class CampGameEntry(db.Model):
    id                   = db.Column(db.Integer, primary_key=True)
    camp_id              = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)
    date                 = db.Column(db.Date, nullable=False)
    slot_ids             = db.Column(db.String(200), nullable=False)
    time_from            = db.Column(db.String(5))
    time_to              = db.Column(db.String(5))
    title                = db.Column(db.String(200))
    cancels_troop_program= db.Column(db.Boolean, default=False)
    physical_intensity   = db.Column(db.String(20))
    entry_type           = db.Column(db.String(20), nullable=False, default='hra')
    except_troop_ids     = db.Column(db.String(200), nullable=True)  # JSON [troop_id, ...]
    teepee               = db.Column(db.Boolean, default=False, nullable=True)

    def get_slot_ids(self):
        return json.loads(self.slot_ids) if self.slot_ids else []

    def get_except_troop_ids(self):
        if not self.except_troop_ids:
            return set()
        return set(json.loads(self.except_troop_ids))


class ServiceEntry(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    camp_id     = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)
    date        = db.Column(db.Date, nullable=False)
    slot_number = db.Column(db.Integer, nullable=False)
    troop_id    = db.Column(db.Integer, db.ForeignKey('troop.id'), nullable=True)
    troop       = db.relationship('Troop')


class MealEntry(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    camp_id     = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)
    date        = db.Column(db.Date, nullable=False)
    meal_type   = db.Column(db.String(30), nullable=False)
    description = db.Column(db.String(300))


class DayAdvisor(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    camp_id     = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)
    date        = db.Column(db.Date, nullable=False)
    slot_number = db.Column(db.Integer, nullable=False, default=1)
    name        = db.Column(db.String(100))


class ChallengeEntry(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    camp_id   = db.Column(db.Integer, db.ForeignKey('camp.id'), nullable=False)
    date      = db.Column(db.Date, nullable=False)
    challenge = db.Column(db.String(30), nullable=False)  # orli_pera, kapky_rosy, bile_tesaky
    area      = db.Column(db.String(20), nullable=False)  # sever, zapad, jih
    time_from = db.Column(db.String(5), nullable=False)
    time_to   = db.Column(db.String(5), nullable=False)


# ── Pomocné funkce ───────────────────────────────────────────────────────────

def text_color_for_bg(hex_color):
    """Vrátí '#222' nebo '#fff' podle jasu pozadí."""
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return '#222' if luminance > 160 else '#fff'

app.jinja_env.globals['text_color']        = text_color_for_bg
app.jinja_env.globals['CHALLENGE_NAMES']   = CHALLENGE_NAMES
app.jinja_env.globals['CHALLENGE_AREAS']   = CHALLENGE_AREAS
app.jinja_env.globals['CHALLENGE_COLORS']  = CHALLENGE_COLORS
app.jinja_env.globals['AREA_ICONS']        = AREA_ICONS
app.jinja_env.globals['INTENSITY_ICONS']   = INTENSITY_ICONS
app.jinja_env.globals['GAME_TYPE_LABELS']  = GAME_TYPE_LABELS
app.jinja_env.globals['GAME_TYPE_ICONS']   = GAME_TYPE_ICONS
app.jinja_env.globals['GAME_TYPE_COLORS']  = GAME_TYPE_COLORS

def lighten_color(hex_color, factor=0.25):
    hex_color = hex_color.lstrip('#')
    r = int(int(hex_color[0:2], 16) * factor + 255 * (1 - factor))
    g = int(int(hex_color[2:4], 16) * factor + 255 * (1 - factor))
    b = int(int(hex_color[4:6], 16) * factor + 255 * (1 - factor))
    return f'{r:02X}{g:02X}{b:02X}'

def safe_sheet_name(name):
    for ch in r'\/:?*[]':
        name = name.replace(ch, '')
    return name[:31]

def times_overlap(a_from, a_to, b_from, b_to, next_day=False):
    """Vrací True, pokud se časové rozmezí [a_from, a_to) překrývá s [b_from, b_to).
    Podporuje přesnoční rozsahy (kde from > to, např. 22:00–07:00).
    next_day=True: b je z předchozího dne a pokračuje přes půlnoc — kontroluj jen část [00:00, b_to].
    """
    # Normalize '00:00' end → '24:00' only for non-overnight slots
    def norm(t): return '24:00' if t == '00:00' else t

    a_overnight = a_from > a_to and a_to != '00:00'
    b_overnight = b_from > b_to and b_to != '00:00'

    # Efektivní rozsah b (challenge)
    if b_overnight:
        if next_day:
            eff_b_from, eff_b_to = '00:00', b_to   # pokračuje od půlnoci do b_to
        else:
            eff_b_from, eff_b_to = b_from, '24:00'  # běží od b_from do půlnoci
    else:
        eff_b_from, eff_b_to = b_from, norm(b_to)

    # Efektivní rozsah a (slot nebo nová výzva)
    if a_overnight:
        eff_a_from, eff_a_to = a_from, '24:00'      # uvažuj jen část před půlnocí
    else:
        eff_a_from, eff_a_to = a_from, norm(a_to)

    return eff_a_from < eff_b_to and eff_a_to > eff_b_from


def build_slot_challenge_map(camp_id, day, slots):
    """Vrátí {slot_id: [ChallengeEntry]}.
    Zahrnuje výzvy z aktuálního dne + přesnoční výzvy z předchozího dne."""
    today_chs = ChallengeEntry.query.filter_by(camp_id=camp_id, date=day).all()
    prev_date  = day - timedelta(days=1)
    prev_chs   = ChallengeEntry.query.filter_by(camp_id=camp_id, date=prev_date).all()
    overnight  = [c for c in prev_chs if c.time_from > c.time_to]  # přes půlnoc

    result = {}
    for slot in slots:
        same = [c for c in today_chs
                if times_overlap(slot.time_from, slot.time_to, c.time_from, c.time_to)]
        cont = [c for c in overnight
                if times_overlap(slot.time_from, slot.time_to, c.time_from, c.time_to,
                                 next_day=True)]
        result[slot.id] = same + cont
    return result


def is_night_activity(time_to, time_from=None, regular_end=None):
    """True pokud aktivita je noční:
    – končí přesně o půlnoci (00:00),
    – přesahuje přes půlnoc (time_to < time_from, např. 23:00→01:00),
    – nebo končí po konci posledního regulárního slotu (regular_end ≥ '06:00')."""
    if not time_to:
        return False
    if time_to == '00:00':
        return True
    if time_from and time_to < time_from:      # křižuje půlnoc, např. 23:00→01:00
        return True
    # regular_end je použit jen pokud jde o denní čas (≥ 06:00).
    # Pokud by byl "01:00" (přes-půlnoční slot), způsobil by falešné poplachy
    # protože "16:00" > "01:00" = True v string porovnání.
    if regular_end and regular_end >= '06:00' and time_to > regular_end:
        return True
    return False

def game_is_night(game, regular_end=None):
    """True pokud hra/aktivita je noční."""
    return is_night_activity(game.time_to, game.time_from, regular_end)

app.jinja_env.globals['game_is_night']      = game_is_night
app.jinja_env.globals['is_night_activity']  = is_night_activity

_DN_LOG = ['Po', 'Út', 'St', 'Čt', 'Pá', 'So', 'Ne']

def log_activity(camp_id, category, description, link_url=None):
    """Zaznamená aktivitu. Selhání logu nesmí přerušit hlavní akci."""
    try:
        db.session.add(ActivityLog(
            camp_id=camp_id, timestamp=datetime.now(),
            category=category, description=description, link_url=link_url))
        db.session.commit()
    except Exception:
        db.session.rollback()


def last_regular_slot_end(slots):
    """Vrátí time_to posledního 'regulárního' slotu (nepřes-půlnočního).
    Noční slot má time_to < time_from (např. 21:30→01:00), je vyřazen."""
    regular = [s for s in slots if s.time_to > s.time_from]
    if regular:
        return regular[-1].time_to
    return slots[-1].time_to if slots else '21:00'


def is_logged_in():
    return session.get('logged_in', False)

def is_admin():
    return session.get('is_admin', False)

def require_login():
    if not is_logged_in():
        return redirect(url_for('login'))

def require_admin():
    if not is_admin():
        flash('Přístup pouze pro administrátora.', 'danger')
        return redirect(url_for('index'))


# ── Auth ─────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pwd = request.form.get('password', '')
        if pwd == 'Havran.2026':
            session['logged_in'] = True
            session['is_admin']  = True
            return redirect(url_for('index'))
        elif pwd == 'havran':
            session['logged_in'] = True
            session['is_admin']  = False
            return redirect(url_for('index'))
        else:
            flash('Špatné heslo.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── Hlavní stránka ───────────────────────────────────────────────────────────

@app.route('/')
def index():
    r = require_login()
    if r: return r
    camps = Camp.query.order_by(Camp.date_from.desc()).all()
    camp_id = session.get('selected_camp_id')
    if not camp_id and camps:
        camp_id = camps[0].id
        session['selected_camp_id'] = camp_id
    camp = Camp.query.get(camp_id) if camp_id else None

    # Kalendářní mřížka dnů (sloupce = týdny, řádky = Po–Ne)
    calendar_cols = []
    day_index_map = {}
    if camp:
        days = camp.get_days()
        day_index_map = {d: i + 1 for i, d in enumerate(days)}
        if days:
            camp_days_set = set(days)
            first_monday = days[0] - timedelta(days=days[0].weekday())
            last_day     = days[-1]
            last_sunday  = last_day + timedelta(days=(6 - last_day.weekday()) % 7)
            cur = first_monday
            while cur <= last_sunday:
                col = [cur + timedelta(days=i) for i in range(7)]
                calendar_cols.append([d if d in camp_days_set else None for d in col])
                cur += timedelta(days=7)

    activity_log = []
    if camp:
        activity_log = (ActivityLog.query
                        .filter_by(camp_id=camp.id)
                        .order_by(ActivityLog.timestamp.desc())
                        .limit(10).all())

    return render_template('index.html', camps=camps, camp=camp,
                           calendar_cols=calendar_cols, day_index_map=day_index_map,
                           activity_log=activity_log)


@app.route('/troop_program/<int:camp_id>/<int:troop_id>')
def troop_program(camp_id, troop_id):
    r = require_login()
    if r: return r
    camp  = Camp.query.get_or_404(camp_id)
    troop = Troop.query.get_or_404(troop_id)
    slots = SlotTemplate.query.filter_by(camp_id=camp_id).order_by(SlotTemplate.order).all()
    svc_templates = (ServiceSlotTemplate.query
                     .filter_by(camp_id=camp_id)
                     .order_by(ServiceSlotTemplate.slot_number).all())

    days_data = []
    for day in camp.get_days():
        entries = ProgramEntry.query.filter_by(
            camp_id=camp_id, troop_id=troop_id, date=day).all()

        # {slot_id: entry}
        slot_entry_map = {}
        for e in entries:
            for sid in e.get_slot_ids():
                slot_entry_map[sid] = e

        # Táborová hra
        game_entries = CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all()
        game_slot_map = {}
        for g in game_entries:
            for sid in g.get_slot_ids():
                game_slot_map[sid] = g
        cancelled_slots = {sid for g in game_entries if g.cancels_troop_program
                           for sid in g.get_slot_ids()}
        cancel_exceptions_map = {}
        for g in game_entries:
            if g.cancels_troop_program:
                for sid in g.get_slot_ids():
                    cancel_exceptions_map[sid] = g.get_except_troop_ids()

        # Služba tohoto oddílu — per-slot na základě časového překryvu
        svc_entries_troop = ServiceEntry.query.filter_by(
            camp_id=camp_id, date=day, troop_id=troop_id).all()
        svc_slot_set = {s.slot_number for s in svc_entries_troop}
        advisors = DayAdvisor.query.filter_by(camp_id=camp_id, date=day).all()
        advisor_map = {a.slot_number: a for a in advisors}

        slot_service_map = {}  # slot_id -> {'adv': DayAdvisor or None}
        for slot in slots:
            for st in svc_templates:
                if st.slot_number in svc_slot_set and \
                        times_overlap(st.time_from, st.time_to, slot.time_from, slot.time_to):
                    slot_service_map[slot.id] = {'adv': advisor_map.get(st.slot_number)}
                    break

        slot_challenge_map = build_slot_challenge_map(camp_id, day, slots)

        days_data.append({
            'day': day,
            'slot_entry_map': slot_entry_map,
            'game_slot_map': game_slot_map,
            'cancelled_slots': cancelled_slots,
            'cancel_exceptions_map': cancel_exceptions_map,
            'slot_service_map': slot_service_map,
            'slot_challenge_map': slot_challenge_map,
        })

    # Shared map pro celý tábor: "datum_shared_group_id" -> [troop objects]
    all_shared = (ProgramEntry.query.filter_by(camp_id=camp_id)
                  .filter(ProgramEntry.shared_group_id.isnot(None)).all())
    all_troops_dict = {t.id: t for t in Troop.query.filter_by(camp_id=camp_id).all()}
    shared_map = {}
    for e in all_shared:
        key = f"{e.date.isoformat()}_{e.shared_group_id}"
        grp = shared_map.setdefault(key, [])
        t = all_troops_dict.get(e.troop_id)
        if t and t.id not in [x.id for x in grp]:
            grp.append(t)

    all_items = (db.session.query(ProgramItem)
                 .join(ProgramEntry, ProgramItem.entry_id == ProgramEntry.id)
                 .filter(ProgramEntry.camp_id == camp_id,
                         ProgramEntry.troop_id == troop_id)
                 .order_by(ProgramEntry.date, ProgramItem.id)
                 .all())

    edit_mode         = request.args.get('edit', '0') == '1'
    last_slot_time_to = last_regular_slot_end(slots)
    return render_template('troop_program.html',
        camp=camp, troop=troop, slots=slots, days_data=days_data,
        shared_map=shared_map, INTENSITY_LABELS=INTENSITY_LABELS,
        all_items=all_items, edit_mode=edit_mode,
        last_slot_time_to=last_slot_time_to)


@app.route('/meals_overview/<int:camp_id>')
def meals_overview(camp_id):
    r = require_login()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    days_data = []
    for day in camp.get_days():
        meals = MealEntry.query.filter_by(camp_id=camp_id, date=day).all()
        days_data.append({'day': day, 'meal_map': {m.meal_type: m for m in meals}})
    edit_mode = request.args.get('edit', '0') == '1'
    return render_template('meals_overview.html',
        camp=camp, days_data=days_data,
        MEAL_TYPES=MEAL_TYPES, MEAL_LABELS=MEAL_LABELS, edit_mode=edit_mode)


@app.route('/service_overview/<int:camp_id>')
def service_overview(camp_id):
    r = require_login()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    svc_templates = (ServiceSlotTemplate.query
                     .filter_by(camp_id=camp_id)
                     .order_by(ServiceSlotTemplate.slot_number).all())
    troops_map = {t.id: t for t in Troop.query.filter_by(camp_id=camp_id).all()}
    days_data = []
    for day in camp.get_days():
        svc_entries = ServiceEntry.query.filter_by(camp_id=camp_id, date=day).all()
        _svc_map = {}
        for _s in svc_entries:
            _svc_map.setdefault(_s.slot_number, []).append(_s)
        advisors = DayAdvisor.query.filter_by(camp_id=camp_id, date=day).all()
        days_data.append({
            'day': day,
            'svc_map': _svc_map,
            'advisor_map': {a.slot_number: a for a in advisors},
        })
    edit_mode = request.args.get('edit', '0') == '1'
    return render_template('service_overview.html',
        camp=camp, svc_templates=svc_templates,
        troops_map=troops_map, days_data=days_data, edit_mode=edit_mode)


@app.route('/game_overview/<int:camp_id>')
def game_overview(camp_id):
    r = require_login()
    if r: return r
    camp  = Camp.query.get_or_404(camp_id)
    entry_type_filter = request.args.get('entry_type')  # 'hra', 'aktivita', nebo None
    slots = SlotTemplate.query.filter_by(camp_id=camp_id).order_by(SlotTemplate.order).all()
    days_data = []
    for day in camp.get_days():
        game_entries = CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all()
        if entry_type_filter:
            game_entries = [g for g in game_entries if (g.entry_type or 'hra') == entry_type_filter]
        game_slot_map = {}
        game_primary = {}
        for g in game_entries:
            ids = g.get_slot_ids()
            if ids:
                game_primary[g.id] = ids[0]
            for sid in ids:
                game_slot_map[sid] = g
        slot_challenge_map = build_slot_challenge_map(camp_id, day, slots)
        days_data.append({'day': day, 'game_slot_map': game_slot_map, 'game_primary': game_primary,
                          'slot_challenge_map': slot_challenge_map})
    last_slot_time_to = last_regular_slot_end(slots)
    edit_mode = request.args.get('edit', '0') == '1'
    return render_template('game_overview.html',
        camp=camp, slots=slots, days_data=days_data,
        entry_type_filter=entry_type_filter,
        INTENSITY_LABELS=INTENSITY_LABELS,
        last_slot_time_to=last_slot_time_to, edit_mode=edit_mode)


@app.route('/master_overview/<int:camp_id>')
def master_overview(camp_id):
    r = require_login()
    if r: return r
    camp   = Camp.query.get_or_404(camp_id)
    slots  = SlotTemplate.query.filter_by(camp_id=camp_id).order_by(SlotTemplate.order).all()
    troops = Troop.query.filter_by(camp_id=camp_id).order_by(Troop.order, Troop.name).all()
    troops_dict = {t.id: t for t in troops}
    svc_templates = (ServiceSlotTemplate.query
                     .filter_by(camp_id=camp_id)
                     .order_by(ServiceSlotTemplate.slot_number).all())

    days_data = []
    for day in camp.get_days():
        game_entries = CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all()
        game_slot_map_hra      = {}
        game_slot_map_aktivita = {}
        cancelled_slots = set()
        for g in game_entries:
            for sid in g.get_slot_ids():
                if (g.entry_type or 'hra') == 'aktivita':
                    game_slot_map_aktivita[sid] = g
                else:
                    game_slot_map_hra[sid] = g
                if g.cancels_troop_program:
                    cancelled_slots.add(sid)

        troop_slot_maps = {}
        for troop in troops:
            entries = ProgramEntry.query.filter_by(
                camp_id=camp_id, troop_id=troop.id, date=day).all()
            sem = {}
            for e in entries:
                for sid in e.get_slot_ids():
                    sem[sid] = e
            troop_slot_maps[troop.id] = sem

        # Služby: seznam {template, troops, adv} pro každý slot šablony
        svc_entries_day = ServiceEntry.query.filter_by(camp_id=camp_id, date=day).all()
        svc_by_slotnum  = {}
        for _s in svc_entries_day:
            svc_by_slotnum.setdefault(_s.slot_number, []).append(_s)
        advisors_day    = DayAdvisor.query.filter_by(camp_id=camp_id, date=day).all()
        advisor_map     = {a.slot_number: a for a in advisors_day}
        day_services = []
        for st in svc_templates:
            svc_list = svc_by_slotnum.get(st.slot_number, [])
            troops_for_slot = [troops_dict[_s.troop_id] for _s in svc_list
                               if _s.troop_id and _s.troop_id in troops_dict]
            adv = advisor_map.get(st.slot_number)
            day_services.append({'template': st, 'troops': troops_for_slot, 'adv': adv})

        # Výjimky z rušení programu
        cancel_exceptions_map = {}
        for g in game_entries:
            if g.cancels_troop_program:
                for sid in g.get_slot_ids():
                    cancel_exceptions_map[sid] = g.get_except_troop_ids()

        # Jídelníček
        meals   = MealEntry.query.filter_by(camp_id=camp_id, date=day).all()
        meal_map = {m.meal_type: m for m in meals}

        # Výzvy
        slot_challenge_map = build_slot_challenge_map(camp_id, day, slots)

        # Sloty, kde žádný oddíl nemá vlastní záznam → použít rowspan přes všechny oddíly
        all_troop_entry_slots = set()
        for t in troops:
            all_troop_entry_slots.update(troop_slot_maps[t.id].keys())
        # Sloty s výjimkami nelze sloučit přes rowspan – různé oddíly mají různý obsah
        slots_with_exceptions = {sid for sid, exc in cancel_exceptions_map.items() if exc}
        full_span_cancelled_slots = (cancelled_slots
                                     - all_troop_entry_slots - slots_with_exceptions)
        full_span_akt_slots = (set(game_slot_map_aktivita.keys())
                               - all_troop_entry_slots - cancelled_slots)

        days_data.append({
            'day':                        day,
            'game_slot_map_hra':          game_slot_map_hra,
            'game_slot_map_aktivita':     game_slot_map_aktivita,
            'cancelled_slots':            cancelled_slots,
            'cancel_exceptions_map':      cancel_exceptions_map,
            'troop_slot_maps':            troop_slot_maps,
            'day_services':               day_services,
            'meal_map':                   meal_map,
            'full_span_cancelled_slots':  full_span_cancelled_slots,
            'full_span_akt_slots':        full_span_akt_slots,
            'slot_challenge_map':         slot_challenge_map,
        })

    last_slot_time_to = last_regular_slot_end(slots)
    return render_template('master_overview.html',
        camp=camp, slots=slots, troops=troops, days_data=days_data,
        INTENSITY_LABELS=INTENSITY_LABELS,
        MEAL_LABELS=MEAL_LABELS, MEAL_TYPES=MEAL_TYPES,
        last_slot_time_to=last_slot_time_to)


@app.route('/challenges_overview/<int:camp_id>')
def challenges_overview(camp_id):
    r = require_login()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    edit_mode = request.args.get('edit', '0') == '1'
    days_data = []
    for day in camp.get_days():
        entries = (ChallengeEntry.query
                   .filter_by(camp_id=camp_id, date=day)
                   .order_by(ChallengeEntry.area, ChallengeEntry.time_from)
                   .all())
        if entries or edit_mode:
            days_data.append({'day': day, 'entries': entries})
    return render_template('challenges_overview.html',
        camp=camp, days_data=days_data, edit_mode=edit_mode,
        CHALLENGE_NAMES=CHALLENGE_NAMES, CHALLENGE_AREAS=CHALLENGE_AREAS,
        CHALLENGE_COLORS=CHALLENGE_COLORS, AREA_ICONS=AREA_ICONS)


@app.route('/teepee_overview/<int:camp_id>')
def teepee_overview(camp_id):
    r = require_login()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    troops_dict = {t.id: t for t in Troop.query.filter_by(camp_id=camp_id).all()}
    reservations = []
    for day in camp.get_days():
        day_res = []
        # Programové záznamy oddílů
        for _pe in (ProgramEntry.query
                    .filter_by(camp_id=camp_id, date=day)
                    .filter(ProgramEntry.teepee == True).all()):
            _t = troops_dict.get(_pe.troop_id)
            day_res.append({
                'time_from': _pe.time_from or '',
                'time_to':   _pe.time_to   or '',
                'troop':     _t,
                'label':     _pe.title or '—',
                'type':      'program',
            })
        # Celotáborové aktivity / táborová hra
        for _cge in (CampGameEntry.query
                     .filter_by(camp_id=camp_id, date=day)
                     .filter(CampGameEntry.teepee == True).all()):
            day_res.append({
                'time_from': _cge.time_from or '',
                'time_to':   _cge.time_to   or '',
                'troop':     None,
                'label':     _cge.title or '—',
                'type':      _cge.entry_type or 'hra',
            })
        day_res.sort(key=lambda x: x['time_from'])
        if day_res:
            reservations.append({'day': day, 'entries': day_res})
    return render_template('teepee_overview.html',
        camp=camp, reservations=reservations)


@app.route('/edit_challenges', methods=['GET', 'POST'])
def edit_challenges():
    r = require_login()
    if r: return r
    camp_id  = request.values.get('camp_id', type=int)
    date_str = request.values.get('date')
    camp = Camp.query.get_or_404(camp_id)
    day  = date.fromisoformat(date_str)

    if request.method == 'POST':
        ch    = request.form.get('challenge')
        area  = request.form.get('area')
        tf    = request.form.get('time_from', '')
        tt    = request.form.get('time_to', '')
        if ch and area and tf and tt:
            existing = ChallengeEntry.query.filter_by(
                camp_id=camp_id, date=day, area=area).all()
            conflict = any(times_overlap(tf, tt, e.time_from, e.time_to)
                           for e in existing)
            if conflict:
                flash(f'V oblasti {CHALLENGE_AREAS[area]} se v tomto čase již koná jiná výzva.', 'danger')
            else:
                db.session.add(ChallengeEntry(
                    camp_id=camp_id, date=day,
                    challenge=ch, area=area, time_from=tf, time_to=tt))
                db.session.commit()
                log_activity(camp_id=camp_id, category='challenge',
                    description=f"Výzva {CHALLENGE_NAMES[ch]}: přidána "
                                f"({AREA_ICONS[area]}, {_DN_LOG[day.weekday()]} {day.strftime('%-d. %-m.')})",
                    link_url=url_for('challenges_overview', camp_id=camp_id))
                flash('Výzva přidána.', 'success')
        _next = request.form.get('next', '') or request.args.get('next', '')
        if _next and _next.startswith('/'):
            return redirect(_next)
        return redirect(url_for('edit_challenges', camp_id=camp_id, date=date_str))

    entries = ChallengeEntry.query.filter_by(camp_id=camp_id, date=day).order_by(
        ChallengeEntry.area, ChallengeEntry.time_from).all()
    next_url = request.args.get('next', '')
    return render_template('edit_challenges.html',
        camp=camp, day=day, entries=entries,
        CHALLENGE_NAMES=CHALLENGE_NAMES, CHALLENGE_AREAS=CHALLENGE_AREAS,
        next_url=next_url)


@app.route('/delete_challenge/<int:entry_id>', methods=['POST'])
def delete_challenge(entry_id):
    r = require_login()
    if r: return r
    e = ChallengeEntry.query.get_or_404(entry_id)
    camp_id, date_str = e.camp_id, e.date.isoformat()
    _ch_desc = (f"Výzva {CHALLENGE_NAMES[e.challenge]}: smazána "
                f"({AREA_ICONS[e.area]}, {_DN_LOG[e.date.weekday()]} {e.date.strftime('%-d. %-m.')})")
    db.session.delete(e)
    db.session.commit()
    log_activity(camp_id=e.camp_id, category='challenge', description=_ch_desc,
                 link_url=url_for('challenges_overview', camp_id=e.camp_id))
    flash('Výzva smazána.', 'info')
    _next = request.args.get('next', '')
    if _next and _next.startswith('/'):
        return redirect(_next)
    return redirect(url_for('edit_challenges', camp_id=camp_id, date=date_str))


@app.route('/export_excel/<int:camp_id>')
def export_excel(camp_id):
    r = require_login()
    if r: return r
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('Knihovna openpyxl není dostupná. Spusť: pip install openpyxl', 'danger')
        return redirect(url_for('index'))

    camp = Camp.query.get_or_404(camp_id)
    slots = SlotTemplate.query.filter_by(camp_id=camp_id).order_by(SlotTemplate.order).all()
    troops = Troop.query.filter_by(camp_id=camp_id).order_by(Troop.order, Troop.name).all()
    svc_templates = (ServiceSlotTemplate.query.filter_by(camp_id=camp_id)
                     .order_by(ServiceSlotTemplate.slot_number).all())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    DN_FULL  = ['Pondělí','Úterý','Středa','Čtvrtek','Pátek','Sobota','Neděle']
    DN_SHORT = ['Po','Út','St','Čt','Pá','So','Ne']

    def _h1(ws, row, col, text, bg='1F2937'):
        c = ws.cell(row, col, text)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        return c

    def _h2(ws, row, col, text):
        c = ws.cell(row, col, text)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill('solid', fgColor='E5E7EB')
        c.alignment = Alignment(wrap_text=True, vertical='top')
        return c

    def _c(ws, row, col, text, bg=None):
        c = ws.cell(row, col, text)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
        return c

    def _merge(ws, row, c1, c2):
        if c2 > c1:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

    # ── Listy dle dní ─────────────────────────────────────────────────────────
    for day in camp.get_days():
        ws = wb.create_sheet(
            title=safe_sheet_name(f"{DN_SHORT[day.weekday()]} {day.strftime('%-d.%-m.')}"))
        ncols = len(slots) + 1

        meals     = MealEntry.query.filter_by(camp_id=camp_id, date=day).all()
        meal_map  = {m.meal_type: m for m in meals}
        svc_ents  = ServiceEntry.query.filter_by(camp_id=camp_id, date=day).all()
        svc_map   = {s.slot_number: s for s in svc_ents}
        advisors  = DayAdvisor.query.filter_by(camp_id=camp_id, date=day).all()
        adv_map   = {a.slot_number: a for a in advisors}
        entries   = ProgramEntry.query.filter_by(camp_id=camp_id, date=day).all()
        slot_troop_map = {}
        entry_primary  = {}
        for e in entries:
            ids = e.get_slot_ids()
            if ids: entry_primary[e.id] = ids[0]
            for sid in ids:
                slot_troop_map.setdefault(sid, {})[e.troop_id] = e
        game_entries = CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all()
        game_slot_map = {}
        game_primary  = {}
        for g in game_entries:
            ids = g.get_slot_ids()
            if ids: game_primary[g.id] = ids[0]
            for sid in ids: game_slot_map[sid] = g
        cancelled = {sid for g in game_entries if g.cancels_troop_program
                     for sid in g.get_slot_ids()}

        r = 1
        c = ws.cell(r, 1, f"{DN_FULL[day.weekday()]} {day.strftime('%-d. %-m. %Y')}")
        c.font = Font(bold=True, size=14)
        _merge(ws, r, 1, ncols); r += 2

        _h1(ws, r, 1, 'JÍDELNÍČEK', bg='92400E')
        _merge(ws, r, 1, 2); r += 1
        for mt in MEAL_TYPES:
            m = meal_map.get(mt)
            _h2(ws, r, 1, MEAL_LABELS[mt])
            _c(ws, r, 2, m.description if m and m.description else '—'); r += 1
        r += 1

        if svc_templates:
            _h1(ws, r, 1, 'SLUŽBA A RÁDCE DNE', bg='6A1B9A')
            _merge(ws, r, 1, 3); r += 1
            for st in svc_templates:
                svc = svc_map.get(st.slot_number)
                adv = adv_map.get(st.slot_number)
                _h2(ws, r, 1, f"Blok {st.slot_number}  {st.time_from}–{st.time_to}")
                _c(ws, r, 2, svc.troop.name if svc and svc.troop else '—')
                _c(ws, r, 3, adv.name if adv and adv.name else '—'); r += 1
            r += 1

        _h1(ws, r, 1, 'Oddíl / Družina')
        for i, slot in enumerate(slots, 2):
            _h1(ws, r, i, f"{slot.name}\n{slot.time_from}–{slot.time_to}")
        r += 1

        if game_slot_map:
            c = ws.cell(r, 1, 'Táborová hra')
            c.font = Font(bold=True, color='1565C0')
            c.alignment = Alignment(vertical='top')
            shown = None
            for i, slot in enumerate(slots, 2):
                g = game_slot_map.get(slot.id)
                if g and g.id != shown:
                    shown = g.id
                    parts = [g.title or '— bez názvu']
                    if g.physical_intensity: parts.append(INTENSITY_LABELS[g.physical_intensity])
                    if g.cancels_troop_program: parts.append('Ruší program oddílů')
                    _c(ws, r, i, '\n'.join(parts), bg='BBDEFB')
                elif g:
                    _c(ws, r, i, '↑', bg='BBDEFB')
            r += 1

        for troop in troops:
            c = ws.cell(r, 1, troop.name)
            c.font = Font(bold=True); c.alignment = Alignment(vertical='top')
            shown = None
            for i, slot in enumerate(slots, 2):
                entry = slot_troop_map.get(slot.id, {}).get(troop.id)
                if entry:
                    lc = lighten_color(troop.color)
                    if entry.id != shown:
                        shown = entry.id
                        esl = [s for s in slots if s.id in entry.get_slot_ids()]
                        t_from = esl[0].time_from if esl else entry.time_from
                        t_to   = esl[-1].time_to  if esl else entry.time_to
                        text = entry.title or '— bez názvu'
                        if t_from and t_to:
                            text += f"\n{t_from}–{t_to}"
                        _c(ws, r, i, text, bg=lc)
                    else:
                        _c(ws, r, i, '↑', bg=lc)
                elif slot.id in cancelled and game_slot_map.get(slot.id):
                    g = game_slot_map[slot.id]
                    _c(ws, r, i, f"Tábor. hra: {g.title or '—'}", bg='BBDEFB')
                else:
                    _c(ws, r, i, '—')
            r += 1

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 18
        for i in range(len(slots)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 24

    # ── Listy dle oddílů ──────────────────────────────────────────────────────
    for troop in troops:
        ws = wb.create_sheet(title=safe_sheet_name(troop.name))
        _h1(ws, 1, 1, 'Den')
        for i, slot in enumerate(slots, 2):
            _h1(ws, 1, i, f"{slot.name}\n{slot.time_from}–{slot.time_to}")

        for ri, day in enumerate(camp.get_days(), 2):
            entries = ProgramEntry.query.filter_by(
                camp_id=camp_id, troop_id=troop.id, date=day).all()
            sem = {}
            for e in entries:
                for sid in e.get_slot_ids(): sem[sid] = e
            game_ents = CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all()
            gsm = {}
            for g in game_ents:
                for sid in g.get_slot_ids(): gsm[sid] = g
            canc = {sid for g in game_ents if g.cancels_troop_program
                    for sid in g.get_slot_ids()}

            c = ws.cell(ri, 1, f"{DN_SHORT[day.weekday()]} {day.strftime('%-d. %-m.')}")
            c.font = Font(bold=True); c.alignment = Alignment(vertical='top')
            shown = None
            for i, slot in enumerate(slots, 2):
                entry = sem.get(slot.id)
                if entry:
                    lc = lighten_color(troop.color)
                    if entry.id != shown:
                        shown = entry.id
                        esl = [s for s in slots if s.id in entry.get_slot_ids()]
                        t_from = esl[0].time_from if esl else entry.time_from
                        t_to   = esl[-1].time_to  if esl else entry.time_to
                        text = entry.title or '—'
                        if t_from and t_to:
                            text += f"\n{t_from}–{t_to}"
                        _c(ws, ri, i, text, bg=lc)
                    else:
                        _c(ws, ri, i, '↑', bg=lc)
                elif slot.id in canc and gsm.get(slot.id):
                    _c(ws, ri, i, f"Tábor. hra: {gsm[slot.id].title or '—'}", bg='BBDEFB')
                else:
                    _c(ws, ri, i, '—')

        ws.column_dimensions['A'].width = 12
        for i in range(len(slots)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 24

    # ── Jídelníček ────────────────────────────────────────────────────────────
    ws = wb.create_sheet(title='Jídelníček')
    _h1(ws, 1, 1, 'Den')
    for i, mt in enumerate(MEAL_TYPES, 2):
        _h1(ws, 1, i, MEAL_LABELS[mt])
    for ri, day in enumerate(camp.get_days(), 2):
        meals = MealEntry.query.filter_by(camp_id=camp_id, date=day).all()
        mm = {m.meal_type: m for m in meals}
        c = ws.cell(ri, 1, f"{DN_SHORT[day.weekday()]} {day.strftime('%-d. %-m.')}")
        c.font = Font(bold=True); c.alignment = Alignment(vertical='top')
        for i, mt in enumerate(MEAL_TYPES, 2):
            m = mm.get(mt)
            _c(ws, ri, i, m.description if m and m.description else '—')
    ws.column_dimensions['A'].width = 12
    for i in range(len(MEAL_TYPES)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 22

    # ── Služby ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet(title='Služby')
    _h1(ws, 1, 1, 'Den')
    for idx, st in enumerate(svc_templates):
        _h1(ws, 1, 2 + idx * 2, f"Blok {st.slot_number}\n{st.time_from}–{st.time_to}\nOddíl")
        _h1(ws, 1, 3 + idx * 2, f"Blok {st.slot_number}\nRádce")
    for ri, day in enumerate(camp.get_days(), 2):
        svc_ents = ServiceEntry.query.filter_by(camp_id=camp_id, date=day).all()
        svc_map  = {s.slot_number: s for s in svc_ents}
        advisors = DayAdvisor.query.filter_by(camp_id=camp_id, date=day).all()
        adv_map  = {a.slot_number: a for a in advisors}
        c = ws.cell(ri, 1, f"{DN_SHORT[day.weekday()]} {day.strftime('%-d. %-m.')}")
        c.font = Font(bold=True); c.alignment = Alignment(vertical='top')
        for idx, st in enumerate(svc_templates):
            svc = svc_map.get(st.slot_number)
            adv = adv_map.get(st.slot_number)
            _c(ws, ri, 2 + idx * 2, svc.troop.name if svc and svc.troop else '—')
            _c(ws, ri, 3 + idx * 2, adv.name if adv and adv.name else '—')
    ws.column_dimensions['A'].width = 12
    for i in range(len(svc_templates) * 2):
        ws.column_dimensions[get_column_letter(i + 2)].width = 18

    # ── Táborová hra ──────────────────────────────────────────────────────────
    ws = wb.create_sheet(title='Táborová hra')
    _h1(ws, 1, 1, 'Den')
    for i, slot in enumerate(slots, 2):
        _h1(ws, 1, i, f"{slot.name}\n{slot.time_from}–{slot.time_to}")
    for ri, day in enumerate(camp.get_days(), 2):
        game_ents = CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all()
        gsm = {}
        gpm = {}
        for g in game_ents:
            ids = g.get_slot_ids()
            if ids: gpm[g.id] = ids[0]
            for sid in ids: gsm[sid] = g
        c = ws.cell(ri, 1, f"{DN_SHORT[day.weekday()]} {day.strftime('%-d. %-m.')}")
        c.font = Font(bold=True); c.alignment = Alignment(vertical='top')
        shown = None
        for i, slot in enumerate(slots, 2):
            g = gsm.get(slot.id)
            if g and g.id != shown:
                shown = g.id
                parts = [g.title or '— bez názvu']
                if g.physical_intensity: parts.append(INTENSITY_LABELS[g.physical_intensity])
                if g.cancels_troop_program: parts.append('Ruší program oddílů')
                _c(ws, ri, i, '\n'.join(parts), bg='BBDEFB')
            elif g:
                _c(ws, ri, i, '↑', bg='BBDEFB')
            else:
                _c(ws, ri, i, '—')
    ws.column_dimensions['A'].width = 12
    for i in range(len(slots)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 24

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = safe_sheet_name(camp.name).replace(' ', '_') + '.xlsx'
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)


@app.route('/admin/backup')
def admin_backup():
    r = require_admin()
    if r: return r
    db_path = os.path.join(basedir, 'tabor.db')
    ts = datetime.now().strftime('%Y_%m_%d_%H%M')
    return send_file(db_path, as_attachment=True, download_name=f'{ts}_tabor_backup.db')


@app.route('/admin/restore', methods=['POST'])
def admin_restore():
    r = require_admin()
    if r: return r
    f = request.files.get('backup_file')
    if not f or not f.filename.endswith('.db'):
        flash('Nahraj platný soubor zálohy (.db).', 'danger')
        return redirect(url_for('admin'))
    db.session.remove()
    db.engine.dispose()
    f.save(os.path.join(basedir, 'tabor.db'))
    flash('Záloha byla obnovena.', 'success')
    return redirect(url_for('admin'))


@app.route('/select_camp/<int:camp_id>')
def select_camp(camp_id):
    r = require_login()
    if r: return r
    session['selected_camp_id'] = camp_id
    return redirect(url_for('index'))


# ── Denní program ────────────────────────────────────────────────────────────

@app.route('/program/<int:camp_id>/<string:date_str>')
def program_day(camp_id, date_str):
    r = require_login()
    if r: return r
    camp  = Camp.query.get_or_404(camp_id)
    day   = date.fromisoformat(date_str)
    slots  = SlotTemplate.query.filter_by(camp_id=camp_id).order_by(SlotTemplate.order).all()
    troops = Troop.query.filter_by(camp_id=camp_id).order_by(Troop.order, Troop.name).all()

    # Program entries → {slot_id: {troop_id: entry}}
    entries = ProgramEntry.query.filter_by(camp_id=camp_id, date=day).all()
    slot_troop_map = {}
    entry_primary = {}   # entry.id -> first slot_id (kde se zobrazuje)
    for e in entries:
        ids = e.get_slot_ids()
        if ids:
            entry_primary[e.id] = ids[0]
        for sid in ids:
            slot_troop_map.setdefault(sid, {})[e.troop_id] = e

    # Společný program: {shared_group_id: [troop objects]}
    troops_dict = {t.id: t for t in troops}
    shared_group_troops = {}
    for e in entries:
        if e.shared_group_id:
            grp = shared_group_troops.setdefault(e.shared_group_id, [])
            t = troops_dict.get(e.troop_id)
            if t and t.id not in [x.id for x in grp]:
                grp.append(t)

    # Camp game entries → {slot_id: entry}
    game_entries = CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all()
    game_slot_map = {}
    game_primary  = {}
    for g in game_entries:
        ids = g.get_slot_ids()
        if ids:
            game_primary[g.id] = ids[0]
        for sid in ids:
            game_slot_map[sid] = g

    # Sloty zrušené táborovou hrou
    cancelled_slots = set()
    for g in game_entries:
        if g.cancels_troop_program:
            cancelled_slots.update(g.get_slot_ids())

    # Služba
    svc_templates = (ServiceSlotTemplate.query
                     .filter_by(camp_id=camp_id)
                     .order_by(ServiceSlotTemplate.slot_number).all())
    svc_entries = ServiceEntry.query.filter_by(camp_id=camp_id, date=day).all()
    svc_map = {}
    for _s in svc_entries:
        svc_map.setdefault(_s.slot_number, []).append(_s)

    # Rádce dne {slot_number: advisor}
    advisors    = DayAdvisor.query.filter_by(camp_id=camp_id, date=day).all()
    advisor_map = {a.slot_number: a for a in advisors}

    # Per-slot service info na základě časového překryvu
    slot_service_troop_ids = {}  # slot_id -> set(troop_id)
    slot_service_info = {}       # slot_id -> list of {'svcs': [...], 'adv': ...}
    for slot in slots:
        troop_ids_in_slot = set()
        infos = []
        for st in svc_templates:
            if times_overlap(st.time_from, st.time_to, slot.time_from, slot.time_to):
                svcs = svc_map.get(st.slot_number, [])
                adv  = advisor_map.get(st.slot_number)
                infos.append({'svcs': svcs, 'adv': adv})
                for _svc in svcs:
                    if _svc.troop_id:
                        troop_ids_in_slot.add(_svc.troop_id)
        slot_service_troop_ids[slot.id] = troop_ids_in_slot
        slot_service_info[slot.id] = infos

    # Výjimky z rušení programu: {slot_id: set(troop_id)}
    cancel_exceptions_map = {}
    for g in game_entries:
        if g.cancels_troop_program:
            for sid in g.get_slot_ids():
                cancel_exceptions_map[sid] = g.get_except_troop_ids()

    # Výzvy: {slot_id: [ChallengeEntry]} — včetně přesčasů z předchozího dne
    slot_challenge_map = build_slot_challenge_map(camp_id, day, slots)

    # Jídelníček
    meals    = MealEntry.query.filter_by(camp_id=camp_id, date=day).all()
    meal_map = {m.meal_type: m for m in meals}

    # Navigace
    all_days = camp.get_days()
    idx      = all_days.index(day) if day in all_days else 0
    prev_day = all_days[idx - 1] if idx > 0 else None
    next_day = all_days[idx + 1] if idx < len(all_days) - 1 else None

    # Mapování jídel na sloty (podle pořadí slotu)
    slot_orders = [s.order for s in slots]
    min_order = min(slot_orders) if slot_orders else 1
    # slot s nejnižším order = dopolední
    morning_id    = slots[0].id if slots else None
    afternoon1_id = slots[1].id if len(slots) > 1 else None
    afternoon2_id = slots[2].id if len(slots) > 2 else None

    edit_mode = request.args.get('edit', '0') == '1'

    last_slot_time_to = last_regular_slot_end(slots)

    return render_template('program_day.html',
        camp=camp, day=day, slots=slots, troops=troops,
        slot_troop_map=slot_troop_map, entry_primary=entry_primary,
        shared_group_troops=shared_group_troops,
        game_slot_map=game_slot_map, game_primary=game_primary,
        cancelled_slots=cancelled_slots,
        svc_templates=svc_templates, svc_map=svc_map,
        slot_service_troop_ids=slot_service_troop_ids,
        slot_service_info=slot_service_info,
        cancel_exceptions_map=cancel_exceptions_map,
        slot_challenge_map=slot_challenge_map,
        meal_map=meal_map, advisor_map=advisor_map,
        prev_day=prev_day, next_day=next_day,
        MEAL_TYPES=MEAL_TYPES, MEAL_LABELS=MEAL_LABELS,
        INTENSITY_LABELS=INTENSITY_LABELS,
        morning_id=morning_id, afternoon1_id=afternoon1_id,
        afternoon2_id=afternoon2_id,
        edit_mode=edit_mode,
        last_slot_time_to=last_slot_time_to,
    )


# ── Editace záznamu oddílu ───────────────────────────────────────────────────

@app.route('/edit_entry', methods=['GET', 'POST'])
def edit_entry():
    r = require_login()
    if r: return r

    camp_id  = request.values.get('camp_id',  type=int)
    troop_id = request.values.get('troop_id', type=int)
    date_str = request.values.get('date')
    slot_id  = request.values.get('slot_id',  type=int)

    camp  = Camp.query.get_or_404(camp_id)
    troop = Troop.query.get_or_404(troop_id)
    day   = date.fromisoformat(date_str)
    slot  = SlotTemplate.query.get_or_404(slot_id)

    existing = None
    for e in ProgramEntry.query.filter_by(camp_id=camp_id, troop_id=troop_id, date=day).all():
        if slot_id in e.get_slot_ids():
            existing = e
            break

    all_slots    = SlotTemplate.query.filter_by(camp_id=camp_id).order_by(SlotTemplate.order).all()
    other_troops = [t for t in Troop.query.filter_by(camp_id=camp_id).order_by(Troop.order, Troop.name).all() if t.id != troop_id]

    if request.method == 'POST':
        title         = request.form.get('title', '').strip()
        teepee_val    = bool(request.form.get('teepee'))
        merges        = [int(x) for x in request.form.getlist('merge_slots')]
        # Seřadit sloty podle pořadí zobrazení (ne podle ID)
        merged_sorted = [s for s in all_slots if s.id in set([slot_id] + merges)]
        slot_ids      = [s.id for s in merged_sorted]
        # Preferuj čas z formuláře (noční aktivity apod.), jinak vezmi čas slotu
        tf = request.form.get('time_from', '').strip() or (merged_sorted[0].time_from if merged_sorted else slot.time_from)
        tt = request.form.get('time_to',   '').strip() or (merged_sorted[-1].time_to  if merged_sorted else slot.time_to)
        share_with_ids = [int(x) for x in request.form.getlist('share_with_troop_id')]
        all_other_ids  = [t.id for t in other_troops]

        # Smaž staré sdílené záznamy z předchozího sdílení tohoto slotu
        if existing and existing.shared_group_id:
            for e in ProgramEntry.query.filter_by(
                    camp_id=camp_id, date=day,
                    shared_group_id=existing.shared_group_id).all():
                if e.troop_id != troop_id:
                    db.session.delete(e)
            db.session.commit()

        # Smaž přebývající záznamy pokrývající tyto sloty (aktuálního oddílu)
        # – zachovej přitom seznam pomůcek pro přesun na nový záznam
        # POZOR: pomůcky načítáme AŽ PO commitu smazání; kdyby byly v session
        # dřív, SQLAlchemy by se pokusil při DELETE nastavit entry_id=NULL, což
        # selže (nullable=False) a způsobí IntegrityError / 500.
        old_entry_ids = []
        for e in ProgramEntry.query.filter_by(camp_id=camp_id, troop_id=troop_id, date=day).all():
            if set(e.get_slot_ids()) & set(slot_ids):
                old_entry_ids.append(e.id)
                db.session.delete(e)
        db.session.commit()
        saved_items = []
        for old_id in old_entry_ids:
            saved_items.extend(ProgramItem.query.filter_by(entry_id=old_id).all())

        entry = ProgramEntry(
            camp_id=camp_id, troop_id=troop_id, date=day,
            slot_ids=json.dumps(slot_ids), time_from=tf, time_to=tt, title=title,
            teepee=teepee_val)
        db.session.add(entry)
        db.session.commit()

        # Přiradit uložené pomůcky novému záznamu
        for item in saved_items:
            item.entry_id = entry.id
        if saved_items:
            db.session.commit()

        # Sdílení programu s jinými oddíly / družinami
        if share_with_ids and set(all_other_ids).issubset(set(share_with_ids)):
            # Vybrány VŠECHNY oddíly → převést na celotáborovou aktivitu
            # Smaž pomůcky nového záznamu (SQLAlchemy cascade by jinak nastavil entry_id=NULL)
            ProgramItem.query.filter_by(entry_id=entry.id).delete(synchronize_session=False)
            db.session.flush()
            db.session.delete(entry)
            for g in CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all():
                if set(g.get_slot_ids()) & set(slot_ids):
                    db.session.delete(g)
            db.session.commit()
            db.session.add(CampGameEntry(
                camp_id=camp_id, date=day, slot_ids=json.dumps(slot_ids),
                time_from=tf, time_to=tt, title=title,
                entry_type='aktivita', cancels_troop_program=True,
                physical_intensity=None))
            db.session.commit()
            flash('Program byl automaticky převeden na celotáborovou aktivitu '
                  '(ruší individuální program oddílů). Fyzickou náročnost a ostatní '
                  'nastavení uprav v sekci Celotáborový program.', 'success')
        elif share_with_ids:
            warnings, successes = [], []
            for share_id in share_with_ids:
                share_troop = Troop.query.get(share_id)
                if not share_troop:
                    continue
                conflict = next((e for e in ProgramEntry.query.filter_by(
                        camp_id=camp_id, troop_id=share_id, date=day).all()
                    if set(e.get_slot_ids()) & set(slot_ids)), None)
                svc_conf = ServiceEntry.query.filter_by(
                    camp_id=camp_id, date=day, troop_id=share_id).first()
                if conflict:
                    warnings.append(f'"{share_troop.name}" (již má program)')
                elif svc_conf:
                    warnings.append(f'"{share_troop.name}" (na službě)')
                else:
                    for e in ProgramEntry.query.filter_by(
                            camp_id=camp_id, troop_id=share_id, date=day).all():
                        if set(e.get_slot_ids()) & set(slot_ids):
                            db.session.delete(e)
                    db.session.commit()
                    db.session.add(ProgramEntry(
                        camp_id=camp_id, troop_id=share_id, date=day,
                        slot_ids=json.dumps(slot_ids), time_from=tf, time_to=tt,
                        title=title, shared_group_id=entry.id))
                    entry.shared_group_id = entry.id
                    successes.append(share_troop.name)
            db.session.commit()
            if warnings:
                flash(f'Sdílení se nepodařilo pro: {", ".join(warnings)}.', 'warning')
            if successes:
                flash(f'Program uložen a sdílen s: {", ".join(successes)}.', 'success')
            else:
                flash('Program uložen.', 'success')
        else:
            flash('Program uložen.', 'success')

        _action = 'upraven' if existing else 'přidán'
        log_activity(
            camp_id=camp_id, category='program',
            description=(f"{troop.name}: {_action} program"
                        f" [{title or "-"}] ({_DN_LOG[day.weekday()]} {day.strftime("%-d. %-m.")}, {slot.name})"),
            link_url=url_for('troop_program', camp_id=camp_id, troop_id=troop_id))
        _next = request.form.get('next', '')
        if _next and _next.startswith('/'):
            return redirect(_next)
        return redirect(url_for('program_day', camp_id=camp_id, date_str=date_str, edit='1'))

    # Sloty dostupné ke sloučení: jen bezprostředně sousední + již sloučené
    primary_idx = next((i for i, s in enumerate(all_slots) if s.id == slot_id), 0)
    existing_ids = set(existing.get_slot_ids() if existing else [slot_id])
    merged_indices = [i for i, s in enumerate(all_slots) if s.id in existing_ids]
    lo = min(merged_indices) if merged_indices else primary_idx
    hi = max(merged_indices) if merged_indices else primary_idx
    allowed_indices = set(range(lo, hi + 1))           # vše co už je sloučeno
    if lo > 0:               allowed_indices.add(lo - 1)   # slot před prvním
    if hi < len(all_slots) - 1: allowed_indices.add(hi + 1)  # slot za posledním
    mergeable_slots = [s for i, s in enumerate(all_slots)
                       if i in allowed_indices and s.id != slot_id]

    existing_shared_troop_ids = set()
    if existing and existing.shared_group_id:
        shared_entries = ProgramEntry.query.filter_by(
            camp_id=camp_id, date=day,
            shared_group_id=existing.shared_group_id).all()
        existing_shared_troop_ids = {e.troop_id for e in shared_entries if e.troop_id != troop_id}

    items        = sorted(existing.items, key=lambda x: x.id) if existing else []
    edit_item_id = request.args.get('edit_item', type=int)
    next_url     = request.args.get('next', '')

    # ── Tee-pee: zjisti konflikt ─────────────────────────────────────────────
    def _tp_overlap(a_from, a_to, b_from, b_to):
        """Jednoduchý časový překryv [a_from, a_to) vs [b_from, b_to)."""
        if not all([a_from, a_to, b_from, b_to]):
            return False
        a_to_n = '24:00' if a_to == '00:00' else a_to
        b_to_n = '24:00' if b_to == '00:00' else b_to
        return a_from < b_to_n and a_to_n > b_from

    own_teepee = bool(existing and existing.teepee)
    teepee_conflict = None
    if not own_teepee:
        _tf_chk = (existing.time_from if existing else slot.time_from) or ''
        _tt_chk = (existing.time_to   if existing else slot.time_to)   or ''
        if _tf_chk and _tt_chk:
            # Zkontroluj ostatní programové záznamy oddílů
            for _pe in (ProgramEntry.query
                        .filter_by(camp_id=camp_id, date=day)
                        .filter(ProgramEntry.teepee == True).all()):
                if existing and _pe.id == existing.id:
                    continue
                if _tp_overlap(_tf_chk, _tt_chk, _pe.time_from or '', _pe.time_to or ''):
                    _t = Troop.query.get(_pe.troop_id)
                    teepee_conflict = {
                        'who':  _t.name if _t else '?',
                        'what': _pe.title or '—',
                        'time': f"{_pe.time_from or ''}–{_pe.time_to or ''}",
                    }
                    break
            # Zkontroluj celotáborové hry/aktivity
            if not teepee_conflict:
                for _cge in (CampGameEntry.query
                             .filter_by(camp_id=camp_id, date=day)
                             .filter(CampGameEntry.teepee == True).all()):
                    if _tp_overlap(_tf_chk, _tt_chk, _cge.time_from or '', _cge.time_to or ''):
                        teepee_conflict = {
                            'who':  GAME_TYPE_LABELS.get(_cge.entry_type or 'hra', 'Aktivita'),
                            'what': _cge.title or '—',
                            'time': f"{_cge.time_from or ''}–{_cge.time_to or ''}",
                        }
                        break

    return render_template('edit_entry.html',
        camp=camp, troop=troop, day=day, slot=slot,
        existing=existing, all_slots=all_slots,
        mergeable_slots=mergeable_slots, other_troops=other_troops,
        existing_shared_troop_ids=existing_shared_troop_ids,
        items=items, edit_item_id=edit_item_id, next_url=next_url,
        own_teepee=own_teepee, teepee_conflict=teepee_conflict)


@app.route('/delete_entry/<int:entry_id>')
def delete_entry(entry_id):
    r = require_login()
    if r: return r
    e = ProgramEntry.query.get_or_404(entry_id)
    camp_id, date_str = e.camp_id, e.date.isoformat()
    _troop_del = Troop.query.get(e.troop_id)
    _t = e.title or '-'
    _desc_del  = (f"{_troop_del.name if _troop_del else '?'}: smazan program [{_t}]"
                  f" ({_DN_LOG[e.date.weekday()]} {e.date.strftime('%-d. %-m.')})")
    _link_del  = url_for('troop_program', camp_id=camp_id, troop_id=e.troop_id)
    ProgramItem.query.filter_by(entry_id=e.id).delete(synchronize_session=False)
    db.session.delete(e)
    db.session.commit()
    log_activity(camp_id=camp_id, category='program', description=_desc_del, link_url=_link_del)
    flash('Záznam smazán.', 'info')
    _next = request.args.get('next', '')
    if _next and _next.startswith('/'):
        return redirect(_next)
    return redirect(url_for('program_day', camp_id=camp_id, date_str=date_str, edit='1'))


# ── Pomůcky / materiál ───────────────────────────────────────────────────────

def _item_redirect(form):
    """Sestaví redirect zpět na edit_entry z hidden fields formuláře."""
    return redirect(url_for('edit_entry',
        camp_id=form.get('camp_id',  type=int),
        troop_id=form.get('troop_id', type=int),
        date=form.get('date'),
        slot_id=form.get('slot_id',  type=int)))


@app.route('/add_program_item', methods=['POST'])
def add_program_item():
    r = require_login()
    if r: return r
    entry_id = request.form.get('entry_id', type=int)
    name     = request.form.get('name', '').strip()
    camp_id  = request.form.get('camp_id',  type=int)
    if entry_id and name:
        db.session.add(ProgramItem(camp_id=camp_id, entry_id=entry_id, name=name))
        db.session.commit()
        _pe = ProgramEntry.query.get(entry_id)
        if _pe:
            _troop_item = Troop.query.get(_pe.troop_id)
            _tname_item = _troop_item.name if _troop_item else '?'
            log_activity(camp_id=camp_id, category='item',
                description=f"{_tname_item}: pomucka [{name}] pridana k [{_pe.title or '-'}] ({_DN_LOG[_pe.date.weekday()]} {_pe.date.strftime('%-d. %-m.')})",
                link_url=url_for('troop_program', camp_id=camp_id, troop_id=_pe.troop_id))
    return _item_redirect(request.form)


@app.route('/delete_program_item/<int:item_id>', methods=['POST'])
def delete_program_item(item_id):
    r = require_login()
    if r: return r
    item   = ProgramItem.query.get_or_404(item_id)
    _iname = item.name
    _ipe   = ProgramEntry.query.get(item.entry_id)
    _icamp = item.camp_id
    db.session.delete(item)
    db.session.commit()
    if _ipe:
        _troop_del_item = Troop.query.get(_ipe.troop_id)
        _tname_del_item = _troop_del_item.name if _troop_del_item else '?'
        log_activity(camp_id=_icamp, category='item',
            description=f"{_tname_del_item}: pomucka [{_iname}] smazana z [{_ipe.title or '-'}] ({_DN_LOG[_ipe.date.weekday()]} {_ipe.date.strftime('%-d. %-m.')})",
            link_url=url_for('troop_program', camp_id=_icamp, troop_id=_ipe.troop_id))
    return _item_redirect(request.form)


@app.route('/edit_program_item/<int:item_id>', methods=['POST'])
def edit_program_item(item_id):
    r = require_login()
    if r: return r
    item = ProgramItem.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
    return _item_redirect(request.form)


@app.route('/toggle_program_item/<int:item_id>', methods=['POST'])
def toggle_program_item(item_id):
    r = require_login()
    if r: return r
    item     = ProgramItem.query.get_or_404(item_id)
    camp_id  = request.form.get('camp_id',  type=int)
    troop_id = request.form.get('troop_id', type=int)
    item.checked = not item.checked
    db.session.commit()
    return redirect(url_for('troop_program', camp_id=camp_id, troop_id=troop_id))


# ── Táborová hra ─────────────────────────────────────────────────────────────

@app.route('/edit_game_entry', methods=['GET', 'POST'])
def edit_game_entry():
    r = require_login()
    if r: return r

    camp_id  = request.values.get('camp_id', type=int)
    date_str = request.values.get('date')
    slot_id  = request.values.get('slot_id', type=int)
    entry_type = request.values.get('entry_type', 'hra')

    camp = Camp.query.get_or_404(camp_id)
    day  = date.fromisoformat(date_str)
    slot = SlotTemplate.query.get_or_404(slot_id)

    existing = None
    for g in CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all():
        if slot_id in g.get_slot_ids():
            existing = g
            break

    if existing:
        entry_type = existing.entry_type or 'hra'

    all_slots = SlotTemplate.query.filter_by(camp_id=camp_id).order_by(SlotTemplate.order).all()

    if request.method == 'POST':
        title  = request.form.get('title', '').strip()
        tf     = request.form.get('time_from', slot.time_from)
        tt_sub = request.form.get('time_to',   slot.time_to)
        cancels   = 'cancels_troop_program' in request.form
        intensity = request.form.get('physical_intensity', '')
        entry_type = request.form.get('entry_type', 'hra')
        except_ids_list = [int(x) for x in request.form.getlist('except_troop_ids')]
        except_ids_json = json.dumps(except_ids_list) if except_ids_list else None
        if not intensity:
            flash('Vyber fyzickou náročnost.', 'danger')
            troops = Troop.query.filter_by(camp_id=camp_id).order_by(Troop.order, Troop.name).all()
            return render_template('edit_game_entry.html',
                camp=camp, day=day, slot=slot,
                existing=existing, all_slots=all_slots,
                entry_type=entry_type, GAME_TYPE_LABELS=GAME_TYPE_LABELS,
                INTENSITY_LABELS=INTENSITY_LABELS, troops=troops)

        merges = [int(x) for x in request.form.getlist('merge_slots')]
        # Seřadit podle pořadí slotů (ne podle ID)
        merged_ordered = [s for s in all_slots if s.id in set([slot_id] + merges)]
        slot_ids = [s.id for s in merged_ordered]
        # time_to: pokud hra pokrývá víc slotů a uživatel nevyplnil vlastní čas
        # (= odeslaná hodnota odpovídá času konce některého ze slotů),
        # automaticky nastav konec na poslední sloučený slot.
        if len(merged_ordered) > 1:
            slot_time_tos = {s.time_to for s in merged_ordered}
            tt = merged_ordered[-1].time_to if tt_sub in slot_time_tos else tt_sub
        else:
            tt = tt_sub
        # Validace sousednosti – sloty musí tvořit nepřerušenou řadu
        indices = [i for i, s in enumerate(all_slots) if s.id in set(slot_ids)]
        if indices and max(indices) - min(indices) + 1 != len(indices):
            flash('Lze sloučit pouze sousední (přilehlé) sloty.', 'danger')
            return redirect(url_for('program_day', camp_id=camp_id, date_str=date_str, edit='1'))

        for g in CampGameEntry.query.filter_by(camp_id=camp_id, date=day).all():
            if set(g.get_slot_ids()) & set(slot_ids):
                db.session.delete(g)
        db.session.commit()

        db.session.add(CampGameEntry(
            camp_id=camp_id, date=day, slot_ids=json.dumps(slot_ids),
            time_from=tf, time_to=tt, title=title,
            cancels_troop_program=cancels, physical_intensity=intensity,
            entry_type=entry_type, except_troop_ids=except_ids_json))
        db.session.commit()
        flash('Táborová hra uložena.', 'success')
        _tlabel = 'Celotáborová aktivita' if entry_type == 'aktivita' else 'Táborová hra'
        log_activity(
            camp_id=camp_id, category=entry_type,
            description=f"{_tlabel} [{title or '-'}] ulozena ({_DN_LOG[day.weekday()]} {day.strftime('%-d. %-m.')})",
            link_url=url_for('game_overview', camp_id=camp_id, entry_type=entry_type))
        _next = request.form.get('next', '')
        if _next and _next.startswith('/'):
            return redirect(_next)
        return redirect(url_for('program_day', camp_id=camp_id, date_str=date_str, edit='1'))

    primary_idx = next((i for i, s in enumerate(all_slots) if s.id == slot_id), 0)
    g_existing_ids = set(existing.get_slot_ids() if existing else [slot_id])
    g_merged_indices = [i for i, s in enumerate(all_slots) if s.id in g_existing_ids]
    g_lo = min(g_merged_indices) if g_merged_indices else primary_idx
    g_hi = max(g_merged_indices) if g_merged_indices else primary_idx
    g_allowed = set(range(g_lo, g_hi + 1))
    if g_lo > 0: g_allowed.add(g_lo - 1)
    if g_hi < len(all_slots) - 1: g_allowed.add(g_hi + 1)
    mergeable_slots = [s for i, s in enumerate(all_slots)
                       if i in g_allowed and s.id != slot_id]

    troops   = Troop.query.filter_by(camp_id=camp_id).order_by(Troop.order, Troop.name).all()
    next_url = request.args.get('next', '')
    return render_template('edit_game_entry.html',
        camp=camp, day=day, slot=slot,
        existing=existing, all_slots=all_slots,
        mergeable_slots=mergeable_slots,
        entry_type=entry_type, GAME_TYPE_LABELS=GAME_TYPE_LABELS,
        INTENSITY_LABELS=INTENSITY_LABELS, troops=troops, next_url=next_url)


@app.route('/delete_game_entry/<int:entry_id>')
def delete_game_entry(entry_id):
    r = require_login()
    if r: return r
    g = CampGameEntry.query.get_or_404(entry_id)
    camp_id, date_str = g.camp_id, g.date.isoformat()
    _gtlabel = 'Celotáborová aktivita' if (g.entry_type == 'aktivita') else 'Táborová hra'
    _gdesc   = (f"{_gtlabel} [{g.title or '-'}] smazana"
                f" ({_DN_LOG[g.date.weekday()]} {g.date.strftime('%-d. %-m.')})")
    _gcat    = g.entry_type or 'hra'
    db.session.delete(g)
    db.session.commit()
    log_activity(camp_id=camp_id, category=_gcat, description=_gdesc,
                 link_url=url_for('game_overview', camp_id=camp_id, entry_type=_gcat))
    flash('Záznam táborové hry smazán.', 'info')
    _next = request.args.get('next', '')
    if _next and _next.startswith('/'):
        return redirect(_next)
    return redirect(url_for('program_day', camp_id=camp_id, date_str=date_str, edit='1'))


# ── Služba a Rádce dne ───────────────────────────────────────────────────────

@app.route('/edit_service', methods=['GET', 'POST'])
def edit_service():
    r = require_login()
    if r: return r

    camp_id  = request.values.get('camp_id', type=int)
    date_str = request.values.get('date')
    camp     = Camp.query.get_or_404(camp_id)
    day      = date.fromisoformat(date_str)
    troops   = Troop.query.filter_by(camp_id=camp_id).order_by(Troop.order, Troop.name).all()

    svc_templates = (ServiceSlotTemplate.query
                     .filter_by(camp_id=camp_id)
                     .order_by(ServiceSlotTemplate.slot_number).all())
    svc_entries = ServiceEntry.query.filter_by(camp_id=camp_id, date=day).all()
    svc_map = {}
    for _s in svc_entries:
        svc_map.setdefault(_s.slot_number, []).append(_s)
    advisors    = DayAdvisor.query.filter_by(camp_id=camp_id, date=day).all()
    advisor_map = {a.slot_number: a for a in advisors}

    if request.method == 'POST':
        for st in svc_templates:
            tids = [int(x) for x in request.form.getlist(f'service_{st.slot_number}')]
            ServiceEntry.query.filter_by(
                camp_id=camp_id, date=day, slot_number=st.slot_number).delete()
            for tid in tids:
                db.session.add(ServiceEntry(
                    camp_id=camp_id, date=day,
                    slot_number=st.slot_number, troop_id=tid))

            adv_name = request.form.get(f'advisor_{st.slot_number}', '').strip()
            adv = advisor_map.get(st.slot_number)
            if adv:
                adv.name = adv_name
            else:
                db.session.add(DayAdvisor(
                    camp_id=camp_id, date=day,
                    slot_number=st.slot_number, name=adv_name))

        db.session.commit()
        flash('Služba a rádce dne uloženi.', 'success')
        log_activity(camp_id=camp_id, category='service',
            description=f"Služby: uloženy {_DN_LOG[day.weekday()]} {day.strftime('%-d. %-m.')}",
            link_url=url_for('service_overview', camp_id=camp_id))
        _next = request.form.get('next', '')
        if _next and _next.startswith('/'):
            return redirect(_next)
        return redirect(url_for('program_day', camp_id=camp_id, date_str=date_str, edit='1'))

    next_url = request.args.get('next', '')
    return render_template('edit_service.html',
        camp=camp, day=day, troops=troops,
        svc_templates=svc_templates, svc_map=svc_map, advisor_map=advisor_map,
        next_url=next_url)


# ── Jídelníček ───────────────────────────────────────────────────────────────

@app.route('/edit_meals', methods=['GET', 'POST'])
def edit_meals():
    r = require_login()
    if r: return r

    camp_id  = request.values.get('camp_id', type=int)
    date_str = request.values.get('date')
    camp     = Camp.query.get_or_404(camp_id)
    day      = date.fromisoformat(date_str)

    meals    = MealEntry.query.filter_by(camp_id=camp_id, date=day).all()
    meal_map = {m.meal_type: m for m in meals}

    if request.method == 'POST':
        for mt in MEAL_TYPES:
            desc = request.form.get(f'meal_{mt}', '').strip()
            if mt in meal_map:
                meal_map[mt].description = desc
            else:
                db.session.add(MealEntry(
                    camp_id=camp_id, date=day,
                    meal_type=mt, description=desc))
        db.session.commit()
        flash('Jídelníček uložen.', 'success')
        log_activity(camp_id=camp_id, category='meals',
            description=f"Jídelníček: uložen {_DN_LOG[day.weekday()]} {day.strftime('%-d. %-m.')}",
            link_url=url_for('meals_overview', camp_id=camp_id))
        _next = request.form.get('next', '')
        if _next and _next.startswith('/'):
            return redirect(_next)
        return redirect(url_for('program_day', camp_id=camp_id, date_str=date_str, edit='1'))

    next_url = request.args.get('next', '')
    return render_template('edit_meals.html',
        camp=camp, day=day, meal_map=meal_map,
        MEAL_TYPES=MEAL_TYPES, MEAL_LABELS=MEAL_LABELS, next_url=next_url)


# ── Historie změn ────────────────────────────────────────────────────────────

@app.route('/activity_log/<int:camp_id>')
def activity_log_page(camp_id):
    r = require_login()
    if r: return r
    camp    = Camp.query.get_or_404(camp_id)
    page    = request.args.get('page', 1, type=int)
    per_page = 20
    offset  = (page - 1) * per_page
    total   = ActivityLog.query.filter_by(camp_id=camp_id).count()
    entries = (ActivityLog.query
               .filter_by(camp_id=camp_id)
               .order_by(ActivityLog.timestamp.desc())
               .offset(offset).limit(per_page).all())
    return render_template('activity_log.html',
        camp=camp, entries=entries, page=page,
        per_page=per_page, total=total,
        has_prev=(page > 1),
        has_next=(offset + per_page < total))


# ── Nápověda ─────────────────────────────────────────────────────────────────

@app.route('/help')
def help_page():
    r = require_login()
    if r: return r
    return render_template('help.html')


# ── Admin ────────────────────────────────────────────────────────────────────

@app.route('/admin')
def admin():
    r = require_admin()
    if r: return r
    camps = Camp.query.order_by(Camp.date_from.desc()).all()
    return render_template('admin.html', camps=camps)


@app.route('/admin/camp/new', methods=['GET', 'POST'])
def admin_new_camp():
    r = require_admin()
    if r: return r
    existing_camps = Camp.query.order_by(Camp.date_from.desc()).all()
    if request.method == 'POST':
        name          = request.form.get('name', '').strip()
        date_from     = date.fromisoformat(request.form.get('date_from'))
        date_to       = date.fromisoformat(request.form.get('date_to'))
        copy_from_id  = request.form.get('copy_from_camp_id', type=int)
        camp = Camp(name=name, date_from=date_from, date_to=date_to)
        db.session.add(camp)
        db.session.commit()

        source = Camp.query.get(copy_from_id) if copy_from_id else None
        if source:
            # Zkopíruj sloty
            for s in SlotTemplate.query.filter_by(camp_id=source.id).order_by(SlotTemplate.order).all():
                db.session.add(SlotTemplate(
                    name=s.name, time_from=s.time_from, time_to=s.time_to,
                    order=s.order, camp_id=camp.id))
            # Zkopíruj sloty služby
            for s in ServiceSlotTemplate.query.filter_by(camp_id=source.id).all():
                db.session.add(ServiceSlotTemplate(
                    slot_number=s.slot_number, time_from=s.time_from,
                    time_to=s.time_to, camp_id=camp.id))
            # Zkopíruj oddíly / družiny
            for t in Troop.query.filter_by(camp_id=source.id).order_by(Troop.order).all():
                db.session.add(Troop(
                    name=t.name, color=t.color, order=t.order, camp_id=camp.id))
            db.session.commit()
            flash(f'Tábor "{camp.name}" vytvořen – nastavení převzato z "{source.name}".', 'success')
        else:
            # Výchozí sloty
            for n, tf, tt, o in [
                ('Dopolední zaměstnání',     '09:00','12:00', 1),
                ('1. odpolední zaměstnání',  '14:00','16:00', 2),
                ('2. odpolední zaměstnání',  '16:00','18:00', 3),
                ('Večerní zaměstnání',       '19:00','21:00', 4),
            ]:
                db.session.add(SlotTemplate(
                    name=n, time_from=tf, time_to=tt, order=o, camp_id=camp.id))
            db.session.add(ServiceSlotTemplate(
                slot_number=1, time_from='00:00', time_to='14:00', camp_id=camp.id))
            db.session.add(ServiceSlotTemplate(
                slot_number=2, time_from='14:00', time_to='00:00', camp_id=camp.id))
            db.session.commit()
            flash(f'Tábor "{camp.name}" vytvořen.', 'success')
        return redirect(url_for('admin'))
    return render_template('admin_camp_form.html', camp=None, existing_camps=existing_camps)


@app.route('/admin/camp/<int:camp_id>/edit', methods=['GET', 'POST'])
def admin_edit_camp(camp_id):
    r = require_admin()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    if request.method == 'POST':
        camp.name      = request.form.get('name', '').strip()
        camp.date_from = date.fromisoformat(request.form.get('date_from'))
        camp.date_to   = date.fromisoformat(request.form.get('date_to'))
        db.session.commit()
        flash('Tábor upraven.', 'success')
        return redirect(url_for('admin'))
    return render_template('admin_camp_form.html', camp=camp)


@app.route('/admin/camp/<int:camp_id>/delete', methods=['POST'])
def admin_delete_camp(camp_id):
    r = require_admin()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    # Smazat všechny závislé záznamy před smazáním tábora
    # (ProgramEntry a ServiceEntry mají FK na Troop, musí jít první)
    entry_ids = [e.id for e in ProgramEntry.query.filter_by(camp_id=camp_id).all()]
    if entry_ids:
        ProgramItem.query.filter(ProgramItem.entry_id.in_(entry_ids)).delete(synchronize_session=False)
    ProgramEntry.query.filter_by(camp_id=camp_id).delete(synchronize_session=False)
    ServiceEntry.query.filter_by(camp_id=camp_id).delete(synchronize_session=False)
    CampGameEntry.query.filter_by(camp_id=camp_id).delete(synchronize_session=False)
    MealEntry.query.filter_by(camp_id=camp_id).delete(synchronize_session=False)
    DayAdvisor.query.filter_by(camp_id=camp_id).delete(synchronize_session=False)
    ChallengeEntry.query.filter_by(camp_id=camp_id).delete(synchronize_session=False)
    ActivityLog.query.filter_by(camp_id=camp_id).delete(synchronize_session=False)
    db.session.delete(camp)  # cascade smaže Troop, SlotTemplate, ServiceSlotTemplate
    db.session.commit()
    flash('Tábor smazán.', 'info')
    return redirect(url_for('admin'))


@app.route('/admin/camp/<int:camp_id>/troops')
def admin_troops(camp_id):
    r = require_admin()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    return render_template('admin_troops.html',
        camp=camp, PALETTE_COLORS=PALETTE_COLORS)


@app.route('/admin/camp/<int:camp_id>/troop/new', methods=['GET', 'POST'])
def admin_new_troop(camp_id):
    r = require_admin()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    if request.method == 'POST':
        name  = request.form.get('name', '').strip()
        color = request.form.get('color', '#4CAF50')
        order = request.form.get('order', type=int, default=0)
        db.session.add(Troop(name=name, color=color, order=order, camp_id=camp_id))
        db.session.commit()
        flash(f'Oddíl "{name}" přidán.', 'success')
        return redirect(url_for('admin_troops', camp_id=camp_id))
    return render_template('admin_troop_form.html',
        camp=camp, troop=None, PALETTE_COLORS=PALETTE_COLORS)


@app.route('/admin/troop/<int:troop_id>/edit', methods=['GET', 'POST'])
def admin_edit_troop(troop_id):
    r = require_admin()
    if r: return r
    troop = Troop.query.get_or_404(troop_id)
    if request.method == 'POST':
        troop.name  = request.form.get('name', '').strip()
        troop.color = request.form.get('color', '#4CAF50')
        troop.order = request.form.get('order', type=int, default=troop.order)
        db.session.commit()
        flash('Oddíl upraven.', 'success')
        return redirect(url_for('admin_troops', camp_id=troop.camp_id))
    return render_template('admin_troop_form.html',
        camp=troop.camp, troop=troop, PALETTE_COLORS=PALETTE_COLORS)


@app.route('/admin/troop/<int:troop_id>/delete', methods=['POST'])
def admin_delete_troop(troop_id):
    r = require_admin()
    if r: return r
    troop = Troop.query.get_or_404(troop_id)
    camp_id = troop.camp_id
    db.session.delete(troop)
    db.session.commit()
    flash('Oddíl smazán.', 'info')
    return redirect(url_for('admin_troops', camp_id=camp_id))


@app.route('/admin/camp/<int:camp_id>/slots')
def admin_slots(camp_id):
    r = require_admin()
    if r: return r
    camp  = Camp.query.get_or_404(camp_id)
    slots = SlotTemplate.query.filter_by(camp_id=camp_id).order_by(SlotTemplate.order).all()
    svc_slots = (ServiceSlotTemplate.query
                 .filter_by(camp_id=camp_id)
                 .order_by(ServiceSlotTemplate.slot_number).all())
    return render_template('admin_slots.html',
        camp=camp, slots=slots, svc_slots=svc_slots)


@app.route('/admin/camp/<int:camp_id>/slot/new', methods=['GET', 'POST'])
def admin_new_slot(camp_id):
    r = require_admin()
    if r: return r
    camp = Camp.query.get_or_404(camp_id)
    if request.method == 'POST':
        name  = request.form.get('name', '').strip()
        tf    = request.form.get('time_from', '09:00')
        tt    = request.form.get('time_to',   '12:00')
        order = request.form.get('order', type=int, default=1)
        db.session.add(SlotTemplate(
            name=name, time_from=tf, time_to=tt, order=order, camp_id=camp_id))
        db.session.commit()
        flash('Slot přidán.', 'success')
        return redirect(url_for('admin_slots', camp_id=camp_id))
    return render_template('admin_slot_form.html', camp=camp, slot=None)


@app.route('/admin/slot/<int:slot_id>/edit', methods=['GET', 'POST'])
def admin_edit_slot(slot_id):
    r = require_admin()
    if r: return r
    slot = SlotTemplate.query.get_or_404(slot_id)
    if request.method == 'POST':
        slot.name      = request.form.get('name', '').strip()
        slot.time_from = request.form.get('time_from', '09:00')
        slot.time_to   = request.form.get('time_to',   '12:00')
        slot.order     = request.form.get('order', type=int, default=slot.order)
        db.session.commit()
        flash('Slot upraven.', 'success')
        return redirect(url_for('admin_slots', camp_id=slot.camp_id))
    return render_template('admin_slot_form.html', camp=slot.camp, slot=slot)


@app.route('/admin/slot/<int:slot_id>/delete', methods=['POST'])
def admin_delete_slot(slot_id):
    r = require_admin()
    if r: return r
    slot = SlotTemplate.query.get_or_404(slot_id)
    camp_id = slot.camp_id
    db.session.delete(slot)
    db.session.commit()
    flash('Slot smazán.', 'info')
    return redirect(url_for('admin_slots', camp_id=camp_id))


@app.route('/admin/service_slot/<int:slot_id>/edit', methods=['GET', 'POST'])
def admin_edit_service_slot(slot_id):
    r = require_admin()
    if r: return r
    slot = ServiceSlotTemplate.query.get_or_404(slot_id)
    if request.method == 'POST':
        slot.time_from = request.form.get('time_from', '00:00')
        slot.time_to   = request.form.get('time_to',   '12:00')
        db.session.commit()
        flash('Slot služby upraven.', 'success')
        return redirect(url_for('admin_slots', camp_id=slot.camp_id))
    return render_template('admin_service_slot_form.html',
        camp=slot.camp, slot=slot)


# ── Init ─────────────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()   # vytvoří activity_log i ostatní nové tabulky
    # ── Migrace: přidání nových sloupců do existujících tabulek ──────────────
    with db.engine.connect() as _conn:
        _cge_cols = {r[1] for r in _conn.execute(text('PRAGMA table_info(camp_game_entry)'))}
        if 'except_troop_ids' not in _cge_cols:
            _conn.execute(text(
                'ALTER TABLE camp_game_entry ADD COLUMN except_troop_ids VARCHAR(200)'))
            _conn.commit()
        if 'teepee' not in _cge_cols:
            _conn.execute(text(
                'ALTER TABLE camp_game_entry ADD COLUMN teepee BOOLEAN DEFAULT 0'))
            _conn.commit()
        _pe_cols = {r[1] for r in _conn.execute(text('PRAGMA table_info(program_entry)'))}
        if 'teepee' not in _pe_cols:
            _conn.execute(text(
                'ALTER TABLE program_entry ADD COLUMN teepee BOOLEAN DEFAULT 0'))
            _conn.commit()
        _pi_cols = {r[1] for r in _conn.execute(text('PRAGMA table_info(program_item)'))}
        if 'camp_id' not in _pi_cols:
            _conn.execute(text(
                'ALTER TABLE program_item ADD COLUMN camp_id INTEGER REFERENCES camp(id)'))
            _conn.commit()
        if 'checked' not in _pi_cols:
            _conn.execute(text(
                'ALTER TABLE program_item ADD COLUMN checked BOOLEAN NOT NULL DEFAULT 0'))
            _conn.commit()
        _se_cols = {r[1] for r in _conn.execute(text('PRAGMA table_info(service_entry)'))}
        if 'troop_id' not in _se_cols:
            _conn.execute(text(
                'ALTER TABLE service_entry ADD COLUMN troop_id INTEGER REFERENCES troop(id)'))
            _conn.commit()

if __name__ == '__main__':
    app.run(debug=True)
